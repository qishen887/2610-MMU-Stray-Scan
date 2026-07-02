# website.py
import os
from uuid import uuid4
from flask import Flask, render_template, url_for, request, jsonify, redirect, flash, send_from_directory, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text
from datetime import datetime, timedelta, timezone
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
app = Flask(__name__)
app.secret_key = 'mmu'  # 保持和你原本 report_page.py 一致的密钥

# Uploads folder 
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 1. 配置数据库路径 (让程序在本地生成 reports.db)
db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reports.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
class User(db.Model):
    # 必须显式指定表名为你在工具里看到的 'users'
    __tablename__ = 'users'  
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), nullable=True)  # Add this line
    email = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(256), nullable=False)
    
    # 💡 提示：由于无法确认你的 reports.db 里有没有 role 这一列
    # 1. 如果你在查看器右侧能看到 role 列，请保留下面这行
    # 2. 如果你的表里【确实没有】role 列，请把下面这行注释掉
    role = db.Column(db.String(50), nullable=True, default='customer')

    def check_password(self, password):
        return check_password_hash(self.password, password)

def utc_now():
    return datetime.now(timezone.utc) + timedelta(hours=8)  # Adjust if you want a different timezone

# 2. 定义你的动物报告数据库表模型 (AnimalReport)
class AnimalReport(db.Model):
    __tablename__ = 'animals_reports'

    id            = db.Column(db.Integer, primary_key=True)
    animal_type   = db.Column(db.String(50),  nullable=False)   # value from dropdown
    custom_animal = db.Column(db.String(100), nullable=True)    # free-text if "other"
    address       = db.Column(db.String(500), nullable=True)   # typed or reverse-geocoded address
    latitude      = db.Column(db.Float, nullable=True)
    longitude     = db.Column(db.Float, nullable=True)
    quantity      = db.Column(db.Integer,     nullable=False)
    health_status = db.Column(db.String(20),  nullable=False)   # healthy/injured/sick/unknown
    details       = db.Column(db.Text,        nullable=True)
    image         = db.Column(db.String(255), nullable=True)    # stored filename only
    status        = db.Column(db.String(20),  nullable=False, default='pending')  # pending/approved/rejected
    created_at    = db.Column(db.DateTime,    default=utc_now)
    submitted_by_email = db.Column(db.String(120), nullable=True)  # None = guest submission

    def to_dict(self):  #Return a JSON-serialisable dict for API responses.
        return {
            "id":            self.id,
            "animal":        self.custom_animal if self.custom_animal else self.animal_type,
            "animal_type":   self.animal_type,
            "custom_animal": self.custom_animal,
            "location":      self.address or (f"{self.latitude}, {self.longitude}" if self.latitude else "—"),
            "quantity":      self.quantity,
            "health":        self.health_status,
            "details":       self.details,
            "image":         self.image if self.image else None,
            "status":        self.status,
            "created_at":    self.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "submitted_by":  self.submitted_by_email or "Guest",
        }

class VetClinic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    address = db.Column(db.String(300), nullable=False)
    phone = db.Column(db.String(30), nullable=False)
    operating_hours = db.Column(db.Text, nullable=False)
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    google_map_link = db.Column(db.String(500), nullable=True)
    image = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=utc_now)

def ensure_user_username_column():
    inspector = inspect(db.engine)
    if 'users' not in inspector.get_table_names():
        return
    columns = {column['name'] for column in inspector.get_columns('users')}
    if 'username' not in columns:
        with db.engine.begin() as connection:
            connection.execute(text("ALTER TABLE users ADD COLUMN username VARCHAR(150)"))

def ensure_vet_clinic_image_column():
    inspector = inspect(db.engine)
    if 'vet_clinic' not in inspector.get_table_names():
        return
    columns = {column['name'] for column in inspector.get_columns('vet_clinic')}
    if 'image' not in columns:
        with db.engine.begin() as connection:
            connection.execute(text("ALTER TABLE vet_clinic ADD COLUMN image VARCHAR(255)"))
    
# 3. 确保在程序启动时，数据库和表能在本地自动创建好
with app.app_context():
    db.create_all()
    ensure_user_username_column()
    ensure_vet_clinic_image_column()

# Routes
app.secret_key = 'mmu'  # same key as in the login file
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 4. 原有的基础页面路由
@app.route('/')
def home():
    return render_template('homepage.html')

@app.route('/home')
def homepage():
    return render_template('homepage.html')

@app.route('/login', methods=['GET', 'POST'])
def show_login():
    if request.method == 'GET':
        if 'user' in session:
            return redirect(url_for('homepage'))
        return render_template('login.html')

    email = (request.form.get('email') or '').strip().lower()
    password = request.form.get('password') or ''

    user = User.query.filter_by(email=email).first()
    if user and check_password_hash(user.password, password):
        session['user'] = user.email
        session['role'] = user.role or 'customer'
        session['display_name'] = user.username or user.email.split('@')[0]
        if session['role'] == 'admin':
            return redirect(url_for('admin'))
        return redirect(url_for('homepage'))

    flash("Invalid email or password. Please try again.")
    return redirect(url_for('show_login'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('show_login'))

@app.route('/session-info')
def session_info():
    if 'user' in session:
        return jsonify({
            "logged_in": True,
            "email": session['user'],
            "role": session.get('role', 'customer'),
            "display_name": session.get('display_name') or session['user'].split('@')[0],
        })
    return jsonify({"logged_in": False})

@app.route('/register')
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    # 1. Add this check to prevent logged-in users from accessing signup
    if 'user' in session:
        return redirect(url_for('homepage'))
    
    if request.method == 'GET':
        return render_template('signup.html')

    username         = request.form.get('username', '').strip() 
    email            = request.form.get('email')
    password         = request.form.get('password')
    confirm_password = request.form.get('confirm_password')

    if not email or not password:
        flash("Email and password are required.")
        return redirect(url_for('signup'))

    if password != confirm_password:
        flash("Passwords do not match!")
        return redirect(url_for('signup'))

    if len(password) != 8:
        flash("Password must be exactly 8 characters!")
        return redirect(url_for('signup'))

    if email.endswith('@mmu.edu.my'):
        user_role = 'admin'
    elif email.endswith('@student.mmu.edu.my'):
        user_role = 'customer'
    else:
        flash("Please use an official MMU email (@mmu.edu.my or @student.mmu.edu.my)")
        return redirect(url_for('signup'))

    if User.query.filter_by(email=email).first():
        flash("Email already registered. Please login.")
        return redirect(url_for('show_login'))

    username = request.form.get('username', '').strip() or email.split('@')[0]
    new_user = User(
        username = username,
        email    = email,
        password = generate_password_hash(password),
        role     = user_role
    )
    db.session.add(new_user)
    db.session.commit()

    flash("Registration successful! Please login.")
    return redirect(url_for('show_login'))

@app.route('/submit', methods=['POST'])
def submit():  # Receive the form, save the image, write a row to the DB.
    try:
        animal_type   = request.form.get('animalType')
        custom_animal = request.form.get('customAnimal') or None
        address       = request.form.get('address') or None
        latitude      = request.form.get('latitude')
        longitude     = request.form.get('longitude')
        quantity      = request.form.get('quantity')
        health_status = request.form.get('healthStatus')
        details       = request.form.get('details') or None

        # Save uploaded image
        file = request.files.get('img')
        saved_filename = None
        if file and file.filename != '':
            saved_filename = file.filename
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], saved_filename))

        # Insert into database 
        report = AnimalReport(
            animal_type   = animal_type,
            custom_animal = custom_animal,
            address       = address,
            latitude      = float(request.form.get('latitude')) if request.form.get('latitude') else None,
            longitude     = float(request.form.get('longitude')) if request.form.get('longitude') else None,
            quantity      = int(quantity),
            health_status = health_status,
            details       = details,
            image         = saved_filename,
            status        = 'pending',
            submitted_by_email = session.get('user'),  # None if not logged in
        )
        db.session.add(report)
        db.session.commit()

        print(f"[DB] Saved report id={report.id}")

        return jsonify({
            "status":  "success",
            "message": "Report submitted successfully",
            "data":    report.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] {e}")
        return jsonify({"status": "error", "message": str(e)})


# 5. 【关键修改】报告页面路由：同时允许 GET（打开页面）和 POST（提交表单）
@app.route('/report', methods=['GET', 'POST'])
def report():
    if request.method == 'POST':
        try:
            # 从前端 report_page.html 的表单里抓取用户填的数据
            animal_type = request.form.get('animal_type')
            quantity = request.form.get('quantity')
            health_status = request.form.get('health_status')
            lat = request.form.get('latitude')
            lng = request.form.get('longitude')
            address = request.form.get('address')
            
            # 创建一条新的数据库记录（注意检查每行末尾的逗号！）
            new_report = AnimalReport(
                animal_type=animal_type,
                quantity=int(quantity),
                health_status=health_status,
                latitude=float(lat),
                longitude=float(lng),
                address=address
            )
            
            # 保存到 SQLite 数据库文件中
            db.session.add(new_report)
            db.session.commit()
            
            flash('Report submitted successfully!', 'success')
            return redirect(url_for('home'))  # 提交成功后跳回主页看地图标记
            
        except Exception as e:
            db.session.rollback()
            return f"Database Error: {str(e)}", 500

    # 如果是普通的打开网页（GET 请求），直接渲染 HTML
    return render_template('report_page.html')

@app.route('/vet-clinics')
@app.route('/vets_clinics')
def vet_clinics():
    clinics = VetClinic.query.order_by(VetClinic.name).all()
    return render_template('vets_clinics.html', form_mode=None, edit_clinic=None, clinics=clinics)


@app.route('/admin/vet-clinics/add', methods=['GET', 'POST'])
@app.route('/vet-clinics/add', methods=['GET', 'POST'])
def add_vet_clinic():
    if request.method == 'POST':
        image_file = request.files.get('image')
        image_filename = None
        if image_file and image_file.filename:
            safe_name = secure_filename(image_file.filename)
            image_filename = f"{uuid4().hex}_{safe_name}"
            image_file.save(os.path.join(app.config['UPLOAD_FOLDER'], image_filename))

        clinic = VetClinic(
            name=request.form['name'],
            address=request.form['address'],
            phone=request.form['phone'],
            operating_hours=request.form['operating_hours'],
            latitude=float(request.form['latitude']) if request.form.get('latitude') else None,
            longitude=float(request.form.get('longitude')) if request.form.get('longitude') else None,
            google_map_link=request.form.get('google_map_link'),
            image=image_filename
        )
        db.session.add(clinic)
        db.session.commit()
        flash('Vet clinic added successfully!', 'success')
        return redirect(url_for('vet_clinics'))
    return render_template('vets_clinics.html', form_mode='add', edit_clinic=None, clinics=VetClinic.query.order_by(VetClinic.name).all())


@app.route('/admin/vet-clinics/edit/<int:clinic_id>', methods=['GET', 'POST'])
@app.route('/vet-clinics/<int:clinic_id>/edit', methods=['GET', 'POST'])
def edit_vet_clinic(clinic_id):
    clinic = VetClinic.query.get_or_404(clinic_id)
    if request.method == 'POST':
        image_file = request.files.get('image')
        if image_file and image_file.filename:
            safe_name = secure_filename(image_file.filename)
            image_filename = f"{uuid4().hex}_{safe_name}"
            image_file.save(os.path.join(app.config['UPLOAD_FOLDER'], image_filename))
            clinic.image = image_filename

        clinic.name = request.form['name']
        clinic.address = request.form['address']
        clinic.phone = request.form['phone']
        clinic.operating_hours = request.form['operating_hours']
        clinic.latitude = float(request.form['latitude']) if request.form.get('latitude') else None
        clinic.longitude = float(request.form.get('longitude')) if request.form.get('longitude') else None
        clinic.google_map_link = request.form.get('google_map_link')
        db.session.commit()
        flash('Vet clinic updated!', 'success')
        return redirect(url_for('vet_clinics'))
    return render_template('vets_clinics.html', form_mode='edit', edit_clinic=clinic, clinics=VetClinic.query.order_by(VetClinic.name).all())


@app.route('/admin/vet-clinics/delete/<int:clinic_id>', methods=['POST'])
@app.route('/vet-clinics/<int:clinic_id>/delete', methods=['POST'])
def delete_vet_clinic(clinic_id):
    clinic = VetClinic.query.get_or_404(clinic_id)
    db.session.delete(clinic)
    db.session.commit()
    flash('Vet clinic deleted.', 'warning')
    return redirect(url_for('vet_clinics'))

from sqlalchemy import and_

@app.route('/api/get_all_reports', methods=['GET'])
def get_all_reports():
    try:
        # Get the filter query parameter sent from the frontend (e.g., 'cat', 'dog')
        animal_type_query = request.args.get('animal_type', '')

        # Filter the database query if a specific animal type is selected and not 'all'
        if animal_type_query and animal_type_query != 'all':
            reports = AnimalReport.query.filter_by(animal_type=animal_type_query).all()
        else:
            # Fallback to retrieving all reports if no filter or 'all' is requested
            reports = AnimalReport.query.all()

        report_list = []
        for r in reports:
            # Safely generate the image static access link if an image file exists in the database
            img_url = url_for('uploaded_file', filename=r.image) if getattr(r, 'image', None) else None
            
            # Pack database record values into a standard data dictionary format
            report_list.append({
                'id': r.id,
                'lat': r.latitude,
                'lng': r.longitude,
                'address': r.address if r.address else f"{r.latitude}, {r.longitude}",
                'animal_type': r.animal_type,
                'quantity': r.quantity,
                'health_status': r.health_status,
                'image_url': img_url
            })
        return jsonify({'status': 'success', 'data': report_list})
    except Exception as e:
        # Return internal server error message if database tracking fails
        return jsonify({'status': 'error', 'message': str(e)}), 500
    
@app.route('/api/filter_reports', methods=['GET'])
def filter_reports():
    """
    API endpoint to filter animal reports based on user selection.
    Expects URL query parameters like: ?types=dog,cat&healths=healthy,injured
    """
    try:
        # 1. Retrieve query parameters from the frontend request
        types_param = request.args.get('types', '')
        healths_param = request.args.get('healths', '')
        
        # 2. Start with a base query that selects all reports
        query = AnimalReport.query
        
        # 3. Apply the Animal Type filter if the user selected any
        if types_param:
            # Convert comma-separated string into a Python list: ['dog', 'cat']
            type_list = types_param.split(',')
            # Use SQLAlchemy's .in_() to filter rows matching any type in the list
            query = query.filter(AnimalReport.animal_type.in_(type_list))
            
        # 4. Apply the Health Status filter if the user selected any
        if healths_param:
            health_list = healths_param.split(',')
            query = query.filter(AnimalReport.health_status.in_(health_list))
            
        # 5. Execute the query to fetch the filtered results from the database
        filtered_reports = query.all()
        
        # 6. Format the database objects into a list of dictionaries for JSON response
        report_list = []
        for r in filtered_reports:
            img_url = url_for('uploaded_file', filename=r.image) if getattr(r, 'image', None) else None
            report_list.append({
                'id': r.id,
                'lat': r.latitude,
                'lng': r.longitude,
                'address': r.address if r.address else f"{r.latitude}, {r.longitude}",
                'animal_type': r.animal_type,
                'quantity': getattr(r, 'quantity', 1), # Default to 1 if not exist
                'health_status': r.health_status,
                'image_url': img_url
            })
            
        # 7. Return the successful response back to the frontend
        return jsonify({'status': 'success', 'data': report_list})

    except Exception as e:
        # Return error details if something goes wrong
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/reports', methods=['GET'])
def get_reports():
    reports = AnimalReport.query.order_by(AnimalReport.created_at.desc()).all()
    return jsonify({"status": "success", "data": [r.to_dict() for r in reports]})


@app.route('/reports/<int:report_id>', methods=['GET'])
def get_report(report_id):
    report = AnimalReport.query.get_or_404(report_id)
    return jsonify({"status": "success", "data": report.to_dict()})


@app.route('/reports/<int:report_id>', methods=['DELETE'])
def delete_report(report_id):
    report = AnimalReport.query.get_or_404(report_id)
    db.session.delete(report)
    db.session.commit()
    return jsonify({"status": "success", "message": f"Report {report_id} deleted."})


@app.route('/admin')
def admin():
    if 'user' not in session:
        flash("Please log in to access the admin panel.")
        return redirect(url_for('show_login'))
    if session.get('role') != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('homepage'))
    return render_template('admin_page.html')


@app.route('/reports/<int:report_id>/status', methods=['PATCH'])
def update_status(report_id):
    try:
        report = AnimalReport.query.get_or_404(report_id)
        new_status = (request.json or {}).get('status')

        if new_status not in ('pending', 'approved', 'rejected'):
            return jsonify({"status": "error", "message": "Invalid status value."}), 400

        report.status = new_status
        db.session.commit()
        return jsonify({"status": "success", "data": report.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/export/excel')
def export_excel():
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    reports = AnimalReport.query.order_by(AnimalReport.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Animal Reports"

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color="2E4057")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    status_fills = {
        "pending": PatternFill("solid", start_color="FFF3CD"),
        "approved": PatternFill("solid", start_color="D4EDDA"),
        "rejected": PatternFill("solid", start_color="F8D7DA"),
    }
    status_fonts = {
        "pending": Font(color="856404", name="Arial", size=10),
        "approved": Font(color="155724", name="Arial", size=10),
        "rejected": Font(color="721C24", name="Arial", size=10),
    }

    headers = ["ID", "Animal", "Custom Animal", "Location", "Quantity",
               "Health Status", "Status", "Details", "Image", "Submitted At", "Submitted By"]
    col_widths = [6, 16, 16, 24, 10, 14, 12, 36, 28, 22, 24]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22
    row_fill_even = PatternFill("solid", start_color="F8F9FA")

    for row_index, report in enumerate(reports, 2):
        row_data = [
            report.id,
            report.animal_type,
            report.custom_animal or "-",
            report.address or (f"{report.latitude}, {report.longitude}" if report.latitude else "-"),
            report.quantity,
            report.health_status,
            report.status,
            report.details or "-",
            report.image or "-",
            report.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            report.submitted_by_email or "Guest",
        ]
        for col_index, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_index == 8))
            cell.font = Font(name="Arial", size=10)
            if row_index % 2 == 0:
                cell.fill = row_fill_even
            if col_index == 7:
                cell.fill = status_fills.get(report.status, PatternFill())
                cell.font = status_fonts.get(report.status, Font(name="Arial", size=10))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_index == 11 and value == "Guest":
                cell.font = Font(italic=True, color="888888", name="Arial", size=10)

        ws.row_dimensions[row_index].height = 18

    last = len(reports) + 2
    ws.cell(row=last, column=1, value="Total").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=last, column=5, value=f'=SUM(E2:E{last-1})').font = Font(bold=True, name="Arial", size=10)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="animal_reports.xlsx"
    )


@app.route('/export/pdf')
def export_pdf():
    import io
    from flask import send_file
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    reports = AnimalReport.query.order_by(AnimalReport.created_at.desc()).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"],
        fontSize=16, spaceAfter=4, textColor=colors.HexColor("#2E4057")
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey, spaceAfter=14
    )
    wrap_style = ParagraphStyle("wrap", parent=styles["Normal"], fontSize=8, leading=10)

    story = [
        Paragraph("Animal Report Export", title_style),
        Paragraph(f"Generated: {utc_now().strftime('%Y-%m-%d %H:%M')} UTC+8 | Total records: {len(reports)}", sub_style),
    ]

    data = [["ID", "Animal", "Location", "Qty", "Health", "Status", "Details", "Submitted", "Submitted By"]]
    for report in reports:
        data.append([
            str(report.id),
            report.custom_animal or report.animal_type,
            Paragraph(report.address or (f"{report.latitude}, {report.longitude}" if report.latitude else "-"), wrap_style),
            str(report.quantity),
            report.health_status.capitalize(),
            report.status.capitalize(),
            Paragraph(report.details or "-", wrap_style),
            report.created_at.strftime("%Y-%m-%d\n%H:%M"),
            Paragraph(report.submitted_by_email or "<i>Guest</i>", wrap_style),
        ])

    table = Table(
        data,
        colWidths=[10 * mm, 22 * mm, 48 * mm, 10 * mm, 16 * mm, 16 * mm, 60 * mm, 25 * mm, 32 * mm],
        repeatRows=1
    )
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4057")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (3, 1), (5, -1), "CENTER"),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]

    status_colors = {
        "pending": colors.HexColor("#FFF3CD"),
        "approved": colors.HexColor("#D4EDDA"),
        "rejected": colors.HexColor("#F8D7DA"),
    }
    for row_index, report in enumerate(reports, 1):
        table_style.append(("BACKGROUND", (5, row_index), (5, row_index), status_colors.get(report.status, colors.white)))

    table.setStyle(TableStyle(table_style))
    story.append(table)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name="animal_reports.pdf")

@app.route('/settings')
def settings_page():
    return render_template('settings_profile.html')

# --- 忘记密码 ---
@app.route('/settings/profile', methods=['POST'])
def update_profile():
    if 'user' not in session:
        flash("Please log in to update your profile.")
        return redirect(url_for('show_login'))

    username = (request.form.get('username') or '').strip()
    if not username:
        flash("Display name cannot be empty.")
        return redirect(url_for('settings_page', tab='profile'))

    user = User.query.filter_by(email=session['user']).first()
    if not user:
        session.clear()
        flash("Session expired. Please log in again.")
        return redirect(url_for('show_login'))

    user.username = username
    db.session.commit()
    session['display_name'] = username
    flash("Profile updated successfully.")
    return redirect(url_for('settings_page', tab='profile'))

@app.route('/settings/password', methods=['POST'])
def change_password():
    if 'user' not in session:
        flash("Please log in to change your password.")
        return redirect(url_for('show_login'))

    current_password = request.form.get('current_password') or ''
    new_password = request.form.get('new_password') or ''
    confirm_password = request.form.get('confirm_password') or ''

    user = User.query.filter_by(email=session['user']).first()
    if not user:
        session.clear()
        flash("Session expired. Please log in again.")
        return redirect(url_for('show_login'))

    if not check_password_hash(user.password, current_password):
        flash("Current password is incorrect.")
        return redirect(url_for('settings_page', tab='password'))

    if new_password != confirm_password:
        flash("New passwords do not match.")
        return redirect(url_for('settings_page', tab='password'))

    if not (len(new_password) == 8 and new_password.isdigit()):
        flash("Password must be exactly 8 digits.")
        return redirect(url_for('settings_page', tab='password'))

    user.password = generate_password_hash(new_password)
    db.session.commit()
    flash("Password updated successfully.")
    return redirect(url_for('settings_page', tab='password'))

@app.route('/forgot-password')
def forgot_password_page():
    return render_template('forgot_password.html')

@app.route('/reset-password', methods=['POST'])
def reset_password():
    email = request.form.get('email').lower().strip()
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')

    # =================【核心修复 7：从数据库查询并修改密码】=================
    user = User.query.filter_by(email=email).first()
    if not user:
        flash("Email address not found. Please register first.")
        return redirect(url_for('forgot_password_page'))

    if password != confirm_password:
        flash("Passwords do not match!")
        return redirect(url_for('forgot_password_page'))
    
    if not (len(password) == 8 and password.isdigit()):
        flash("Format error: Password must be exactly 8 digits!")
        return redirect(url_for('forgot_password_page'))

    # 修改密码并保存
    user.password = generate_password_hash(password)
    db.session.commit()

    flash("Password reset successfully! Please login with your new password.")
    return redirect(url_for('show_login'))
from flask import send_from_directory, url_for

if __name__ == '__main__':
    app.run(debug=True)
