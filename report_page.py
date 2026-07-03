from flask import Flask, redirect, render_template, request, jsonify, send_from_directory, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import inspect, text
from datetime import datetime, timedelta, timezone
import os
import urllib.request
import urllib.parse
import json
from uuid import uuid4
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename


def reverse_geocode(lat, lon):
    if lat is None or lon is None:
        return "—"
    try:
        params = urllib.parse.urlencode({"lat": lat, "lon": lon, "format": "json"})
        url = f"https://nominatim.openstreetmap.org/reverse?{params}"
        req = urllib.request.Request(url, headers={"User-Agent": "AnimalReportApp/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        print(f"[GEOCODE] lat={lat} lon={lon} -> {data}")
        return data.get("display_name", f"{lat}, {lon}")
    except Exception as e:
        print(f"[GEOCODE ERROR] lat={lat} lon={lon} -> {e}")
        return f"{lat}, {lon}"

app = Flask(__name__)
app.secret_key = 'mmu'

# Uploads folder 
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# SQLite database (saved next to report.py as reports.db)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(BASE_DIR, 'reports.db')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

def utc_now():
    return datetime.now(timezone.utc) + timedelta(hours=8)  # Adjust if you want a different timezone

# Model
class AnimalReport(db.Model):
    __tablename__ = 'animals_reports'

    id            = db.Column(db.Integer, primary_key=True)
    animal_type   = db.Column(db.String(50),  nullable=False)   # value from dropdown
    custom_animal = db.Column(db.String(100), nullable=True)    # free-text if "other"
    address       = db.Column(db.String(500), nullable=True)   # typed or reverse-geocoded address
    latitude      = db.Column(db.Float, nullable=True)
    longitude     = db.Column(db.Float, nullable=True)
    quantity      = db.Column(db.Integer,     nullable=False)
    health_status = db.Column(db.String(20),  nullable=False)   # healthy/injured/sick/unknown
    details       = db.Column(db.Text,        nullable=True)
    image         = db.Column(db.String(255), nullable=True)    # stored filename only
    status        = db.Column(db.String(20),  nullable=False, default='pending')  # pending/approved/rejected
    created_at    = db.Column(db.DateTime,    default=utc_now)
    submitted_by_email = db.Column(db.String(120), nullable=True)  # None = guest submission

    def to_dict(self):  #Return a JSON-serialisable dict for API responses.
        return {
            "id":            self.id,
            "animal":        self.custom_animal if self.custom_animal else self.animal_type,
            "animal_type":   self.animal_type,
            "custom_animal": self.custom_animal,
            "location":      self.address or (f"{self.latitude}, {self.longitude}" if self.latitude else "—"),
            "quantity":      self.quantity,
            "health":        self.health_status,
            "details":       self.details,
            "image":         self.image if self.image else None,
            "status":        self.status,
            "created_at":    self.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "submitted_by":  self.submitted_by_email or "Guest",
        }

class User(db.Model):
    __tablename__ = 'users'

    id         = db.Column(db.Integer, primary_key=True)
    username   = db.Column(db.String(80), unique=True, nullable=False)
    email      = db.Column(db.String(120), unique=True, nullable=False)
    password   = db.Column(db.String(255), nullable=False)
    role       = db.Column(db.String(20),  nullable=False, default='customer')

    def check_password(self, password):
        return check_password_hash(self.password, password)

class VetClinic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    address = db.Column(db.String(300), nullable=False)
    phone = db.Column(db.String(30), nullable=False)
    operating_hours = db.Column(db.Text, nullable=False)  # Changed to db.Text for multiline
    latitude = db.Column(db.Float, nullable=True)
    longitude = db.Column(db.Float, nullable=True)
    google_map_link = db.Column(db.String(500), nullable=True) # New Google Map Link field
    image = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


def ensure_vet_clinic_image_column():
    inspector = inspect(db.engine)
    columns = {column['name'] for column in inspector.get_columns('vet_clinic')}
    if 'image' not in columns:
        with db.engine.begin() as connection:
            connection.execute(text("ALTER TABLE vet_clinic ADD COLUMN image VARCHAR(255)"))

def ensure_user_created_at_column():
    return

def ensure_user_username_column():
    inspector = inspect(db.engine)
    if 'users' not in inspector.get_table_names():
        return
    columns = {column['name'] for column in inspector.get_columns('users')}
    if 'username' not in columns:
        with db.engine.begin() as connection:
            connection.execute(text("ALTER TABLE users ADD COLUMN username VARCHAR(150)"))

# Create all tables on first run
with app.app_context():
    db.create_all()
    ensure_vet_clinic_image_column()
    ensure_user_created_at_column()
    ensure_user_username_column()


# Routes
app.secret_key = 'mmu'  # same key as in the login file
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/')
def home():
    return redirect(url_for('homepage'))


@app.route('/home')
def homepage():
    return render_template('homepage.html')

@app.route('/report')
def report():
    return render_template('report_page.html')

# FIXED: Added a secondary route decorator so HTML templates asking for 'show_login' won't crash 
@app.route('/login', methods=['GET', 'POST'], endpoint='login_page')
@app.route('/login-alt', methods=['GET', 'POST'], endpoint='show_login')
def show_login():
    if request.method == 'GET':
        if 'user' in session:
            if session['role'] == 'admin':
                return redirect(url_for('admin'))
            return redirect(url_for('homepage'))
        return render_template('login.html')

    # POST — handle login form submission
    email    = request.form.get('email')
    password = request.form.get('password')

    user = User.query.filter_by(email=email).first()

    if user and user.check_password(password):
        session['user'] = user.email
        session['role'] = user.role
        session['display_name'] = user.username or user.email.split('@')[0]
        if user.role == 'admin':
            return redirect(url_for('admin'))
        return redirect(url_for('homepage'))
    else:
        flash("Invalid email or password. Please try again.")
        return redirect(url_for('login_page'))

 
@app.route('/register')
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    # 1. Add this check to prevent logged-in users from accessing signup
    if 'user' in session:
        return redirect(url_for('homepage'))
    
    if request.method == 'GET':
        return render_template('signup.html')

    email            = request.form.get('email')
    password         = request.form.get('password')
    confirm_password = request.form.get('confirm_password')

    if not email or not password:
        flash("Email and password are required.")
        return redirect(url_for('signup'))

    if password != confirm_password:
        flash("Passwords do not match!")
        return redirect(url_for('signup'))

    if len(password) != 8:
        flash("Password must be exactly 8 characters!")
        return redirect(url_for('signup'))

    if email.endswith('@mmu.edu.my'):
        user_role = 'admin'
    elif email.endswith('@student.mmu.edu.my'):
        user_role = 'customer'
    else:
        flash("Please use an official MMU email (@mmu.edu.my or @student.mmu.edu.my)")
        return redirect(url_for('signup'))

    if User.query.filter_by(email=email).first():
        flash("Email already registered. Please login.")
        return redirect(url_for('login_page'))

    username = request.form.get('username', '').strip() or email.split('@')[0]
    new_user = User(
        username = username,
        email    = email,
        password = generate_password_hash(password),
        role     = user_role
    )
    db.session.add(new_user)
    db.session.commit()

    flash("Registration successful! Please login.")
    return redirect(url_for('login_page'))

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/session-info')
def session_info():
    if 'user' in session:
        return jsonify({"logged_in": True, "email": session['user'], "role": session['role']})
    return jsonify({"logged_in": False})

@app.route('/reverse-geocode', methods=['GET'])
def reverse_geocode_endpoint():
    lat = request.args.get('lat')
    lon = request.args.get('lon')
    if not lat or not lon:
        return jsonify({"status": "error", "message": "lat and lon are required"}), 400
    address = reverse_geocode(lat, lon)
    return jsonify({"status": "success", "address": address})

@app.route('/submit', methods=['POST'])
def submit():  
    try:
        animal_type   = request.form.get('animalType')
        custom_animal = request.form.get('customAnimal') or None
        address       = request.form.get('address') or None
        latitude      = request.form.get('latitude')
        longitude     = request.form.get('longitude')
        if not address and latitude and longitude:
            address = reverse_geocode(latitude, longitude)
        quantity      = request.form.get('quantity')
        health_status = request.form.get('healthStatus')
        details       = request.form.get('details') or None

        # Save uploaded image
        file = request.files.get('img')
        saved_filename = None
        if file and file.filename != '':
            saved_filename = file.filename
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], saved_filename))

        # Insert into database 
        report = AnimalReport(
            animal_type   = animal_type,
            custom_animal = custom_animal,
            address       = address,
            latitude      = float(request.form.get('latitude')) if request.form.get('latitude') else None,
            longitude     = float(request.form.get('longitude')) if request.form.get('longitude') else None,
            quantity      = int(quantity),
            health_status = health_status,
            details       = details,
            image         = saved_filename,
            status        = 'pending',
            submitted_by_email = session.get('user'),  # None if not logged in
        )
        db.session.add(report)
        db.session.commit()

        print(f"[DB] Saved report id={report.id}")

        return jsonify({
            "status":  "success",
            "message": "Report submitted successfully",
            "data":    report.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] {e}")
        return jsonify({"status": "error", "message": str(e)})

@app.route('/settings')
def settings_page():
    return render_template('settings_profile.html')
from flask import send_from_directory, url_for

@app.route('/settings/profile', methods=['POST'])
def update_profile():
    if 'user' not in session:
        flash("Please log in to update your profile.")
        return redirect(url_for('show_login'))

    username = (request.form.get('username') or '').strip()
    if not username:
        flash("Display name cannot be empty.")
        return redirect(url_for('settings_page', tab='profile'))

    user = User.query.filter_by(email=session['user']).first()
    if not user:
        session.clear()
        flash("Session expired. Please log in again.")
        return redirect(url_for('show_login'))

    user.username = username
    db.session.commit()
    session['display_name'] = username
    flash("Profile updated successfully.")
    return redirect(url_for('settings_page', tab='profile'))

@app.route('/settings/password', methods=['POST'])
def change_password():
    if 'user' not in session:
        flash("Please log in to change your password.")
        return redirect(url_for('show_login'))

    current_password = request.form.get('current_password') or ''
    new_password = request.form.get('new_password') or ''
    confirm_password = request.form.get('confirm_password') or ''

    user = User.query.filter_by(email=session['user']).first()
    if not user:
        session.clear()
        flash("Session expired. Please log in again.")
        return redirect(url_for('show_login'))

    if not check_password_hash(user.password, current_password):
        flash("Current password is incorrect.")
        return redirect(url_for('settings_page', tab='password'))

    if new_password != confirm_password:
        flash("New passwords do not match.")
        return redirect(url_for('settings_page', tab='password'))

    if not (len(new_password) == 8 and new_password.isdigit()):
        flash("Password must be exactly 8 digits.")
        return redirect(url_for('settings_page', tab='password'))

    user.password = generate_password_hash(new_password)
    db.session.commit()
    flash("Password updated successfully.")
    return redirect(url_for('settings_page', tab='password'))

@app.route('/api/get_all_reports', methods=['GET'])
def get_all_reports():
    try:
        animal_type_query = request.args.get('animal_type', '')

        if animal_type_query and animal_type_query != 'all':
            reports = AnimalReport.query.filter_by(animal_type=animal_type_query).all()
        else:
            reports = AnimalReport.query.all()

        report_list = []
        for report in reports:
            img_url = url_for('uploaded_file', filename=report.image) if getattr(report, 'image', None) else None
            report_list.append({
                'id': report.id,
                'lat': report.latitude,
                'lng': report.longitude,
                'address': report.address if report.address else f"{report.latitude}, {report.longitude}",
                'animal_type': report.animal_type,
                'quantity': report.quantity,
                'health_status': report.health_status,
                'image_url': img_url
            })
        return jsonify({'status': 'success', 'data': report_list})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/filter_reports', methods=['GET'])
def filter_reports():
    try:
        types_param = request.args.get('types', '')
        healths_param = request.args.get('healths', '')

        query = AnimalReport.query

        if types_param:
            query = query.filter(AnimalReport.animal_type.in_(types_param.split(',')))

        if healths_param:
            query = query.filter(AnimalReport.health_status.in_(healths_param.split(',')))

        report_list = []
        for report in query.all():
            img_url = url_for('uploaded_file', filename=report.image) if getattr(report, 'image', None) else None
            report_list.append({
                'id': report.id,
                'lat': report.latitude,
                'lng': report.longitude,
                'address': report.address if report.address else f"{report.latitude}, {report.longitude}",
                'animal_type': report.animal_type,
                'quantity': getattr(report, 'quantity', 1),
                'health_status': report.health_status,
                'image_url': img_url
            })

        return jsonify({'status': 'success', 'data': report_list})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Optional read endpoints for admin dashboard to list reports, view details, delete, or update status.

@app.route('/reports', methods=['GET'])
def get_reports(): #Return all reports, newest first.
    reports = AnimalReport.query.order_by(AnimalReport.created_at.desc()).all()
    return jsonify({"status": "success", "data": [r.to_dict() for r in reports]})


@app.route('/reports/<int:report_id>', methods=['GET'])
def get_report(report_id):  #Return a single report by ID.
    report = AnimalReport.query.get_or_404(report_id)
    return jsonify({"status": "success", "data": report.to_dict()})


@app.route('/reports/<int:report_id>', methods=['DELETE'])
def delete_report(report_id):   #Delete a report by ID.
    report = AnimalReport.query.get_or_404(report_id)
    db.session.delete(report)
    db.session.commit()
    return jsonify({"status": "success", "message": f"Report {report_id} deleted."})


@app.route('/admin')
def admin():
    if 'user' not in session:
        flash("Please log in to access the admin panel.")
        return redirect(url_for('login_page'))
    if session.get('role') != 'admin':
        flash("Access denied. Admin privileges required.")
        return redirect(url_for('homepage'))
    return render_template('admin_page.html')


@app.route('/reports/<int:report_id>/status', methods=['PATCH'])
def update_status(report_id):   # Admin: change a report status to approved or rejected.
    try:
        report = AnimalReport.query.get_or_404(report_id)
        new_status = request.json.get('status')

        if new_status not in ('pending', 'approved', 'rejected'):
            return jsonify({"status": "error", "message": "Invalid status value."}), 400

        report.status = new_status
        db.session.commit()
        return jsonify({"status": "success", "data": report.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)})


@app.route('/export/excel')
def export_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
 
    reports = AnimalReport.query.order_by(AnimalReport.created_at.desc()).all()
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Animal Reports"
 
    # Header style
    header_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill   = PatternFill("solid", start_color="2E4057")
    header_align  = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
 
    # Status fill colours
    status_fills = {
        "pending":  PatternFill("solid", start_color="FFF3CD"),
        "approved": PatternFill("solid", start_color="D4EDDA"),
        "rejected": PatternFill("solid", start_color="F8D7DA"),
    }
    status_fonts = {
        "pending":  Font(color="856404",  name="Arial", size=10),
        "approved": Font(color="155724",  name="Arial", size=10),
        "rejected": Font(color="721C24",  name="Arial", size=10),
    }
 
    headers = ["ID", "Animal", "Custom Animal", "Location", "Quantity",
               "Health Status", "Status", "Details", "Image", "Submitted At", "Submitted By"]
    col_widths = [6, 16, 16, 24, 10, 14, 12, 36, 28, 22, 24]
 
    # Write headers
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = thin_border
        ws.column_dimensions[cell.column_letter].width = w
 
    ws.row_dimensions[1].height = 22
 
    # Write data rows
    row_fill_even = PatternFill("solid", start_color="F8F9FA")
    for r_idx, report in enumerate(reports, 2):
        row_data = [
            report.id,
            report.animal_type,
            report.custom_animal or "—",
            report.address or (f"{report.latitude}, {report.longitude}" if report.latitude else "—"),
            report.quantity,
            report.health_status,
            report.status,
            report.details or "—",
            report.image or "—",
            report.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            report.submitted_by_email or "Guest",
        ]
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(c_idx == 8))
            cell.font      = Font(name="Arial", size=10)
            if r_idx % 2 == 0:
                cell.fill = row_fill_even
            # Colour the status cell
            if c_idx == 7:
                s = report.status
                cell.fill = status_fills.get(s, PatternFill())
                cell.font = status_fonts.get(s, Font(name="Arial", size=10))
                cell.alignment = Alignment(horizontal="center", vertical="center")
 
            # Style the "Submitted By" cell
            if c_idx == 11:
                if value == "Guest":
                    cell.font = Font(italic=True, color="888888", name="Arial", size=10)
                else:
                    cell.font = Font(color="2E4057", name="Arial", size=10)

        ws.row_dimensions[r_idx].height = 18
 
    # Summary row at the bottom
    last = len(reports) + 2
    ws.cell(row=last, column=1, value="Total").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=last, column=5, value=f'=SUM(E2:E{last-1})').font = Font(bold=True, name="Arial", size=10)
 
    # Freeze the header row
    ws.freeze_panes = "A2"
 
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="animal_reports.xlsx"
    )
 
 
@app.route('/export/pdf')
def export_pdf():
    import io
    from flask import send_file
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER

    def utc_now():
        return datetime.now(timezone.utc) + timedelta(hours=8)  # Adjust if you want a different timezone
 
    reports = AnimalReport.query.order_by(AnimalReport.created_at.desc()).all()
 
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
 
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 fontSize=16, spaceAfter=4, textColor=colors.HexColor("#2E4057"))
    sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.grey, spaceAfter=14)
    wrap_style  = ParagraphStyle("wrap", parent=styles["Normal"], fontSize=8, leading=10)
 
    story = [
        Paragraph("🐾 Animal Report Export", title_style),
        Paragraph(f"Generated: {utc_now().strftime('%Y-%m-%d %H:%M')} UTC+8  |  Total records: {len(reports)}", sub_style),
    ]
 
    # Table data
    col_headers = ["ID", "Animal", "Location", "Qty", "Health", "Status", "Details", "Submitted", "Submitted By"]
    data = [col_headers]
    for rpt in reports:
        data.append([
            str(rpt.id),
            f"{rpt.custom_animal or rpt.animal_type}",
            Paragraph(rpt.address or (f"{rpt.latitude}, {rpt.longitude}" if rpt.latitude else "—"), wrap_style),
            str(rpt.quantity),
            rpt.health_status.capitalize(),
            rpt.status.capitalize(),
            Paragraph(rpt.details or "—", wrap_style),
            rpt.created_at.strftime("%Y-%m-%d\n%H:%M"),
            Paragraph(rpt.submitted_by_email or "<i>Guest</i>", wrap_style),
        ])
 
    col_widths_pdf = [10*mm, 22*mm, 48*mm, 10*mm, 16*mm, 16*mm, 60*mm, 25*mm, 32*mm]
 
    STATUS_COLORS = {
        "pending":  colors.HexColor("#FFF3CD"),
        "approved": colors.HexColor("#D4EDDA"),
        "rejected": colors.HexColor("#F8D7DA"),
    }
 
    tbl = Table(data, colWidths=col_widths_pdf, repeatRows=1)
    base_style = [
        # Header
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#2E4057")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 8),
        ("ALIGN",       (0,0), (-1,0), "CENTER"),
        ("BOTTOMPADDING",(0,0),(-1,0), 8),
        ("TOPPADDING",  (0,0), (-1,0), 8),
        # Body
        ("FONTNAME",    (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",    (0,1), (-1,-1), 8),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("GRID",        (0,0), (-1,-1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("ALIGN",       (0,1), (0,-1), "CENTER"),   # ID
        ("ALIGN",       (3,1), (3,-1), "CENTER"),   # Qty
        ("ALIGN",       (4,1), (5,-1), "CENTER"),   # Health / Status
        ("TOPPADDING",  (0,1), (-1,-1), 5),
        ("BOTTOMPADDING",(0,1),(-1,-1), 5),
    ]
 
    # Colour status cells per row
    for i, rpt in enumerate(reports, 1):
        bg = STATUS_COLORS.get(rpt.status, colors.white)
        base_style.append(("BACKGROUND", (5, i), (5, i), bg))
 
    tbl.setStyle(TableStyle(base_style))
    story.append(tbl)
 
    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name="animal_reports.pdf")


@app.route('/vet-clinics')
@app.route('/vets_clinics')
def vet_clinics():
    clinics = VetClinic.query.order_by(VetClinic.name).all()
    return render_template('vets_clinics.html', form_mode=None, edit_clinic=None, clinics=clinics)

@app.route('/admin/vet-clinics/add', methods=['GET', 'POST'])
@app.route('/vet-clinics/add', methods=['GET', 'POST'])
def add_vet_clinic():
    if 'user' not in session or session.get('role') != 'admin':
        return redirect(url_for('login_page'))
    if request.method == 'POST':
        image_file = request.files.get('image')
        image_filename = None
        if image_file and image_file.filename:
            safe_name = secure_filename(image_file.filename)
            image_filename = f"{uuid4().hex}_{safe_name}"
            image_file.save(os.path.join(app.config['UPLOAD_FOLDER'], image_filename))

        clinic = VetClinic(
            name=request.form['name'],
            address=request.form['address'],
            phone=request.form['phone'],
            operating_hours=request.form['operating_hours'],
            latitude=float(request.form['latitude']) if request.form.get('latitude') else None,
            longitude=float(request.form.get('longitude')) if request.form.get('longitude') else None,
            google_map_link=request.form.get('google_map_link'),
            image=image_filename
        )
        db.session.add(clinic)
        db.session.commit()
        flash('Vet clinic added successfully!', 'success')
        return redirect(url_for('vet_clinics'))
    return render_template('vets_clinics.html', form_mode='add', edit_clinic=None, clinics=VetClinic.query.all())



@app.route('/admin/vet-clinics/edit/<int:clinic_id>', methods=['GET', 'POST'])
@app.route('/vet-clinics/<int:clinic_id>/edit', methods=['GET', 'POST'])
def edit_vet_clinic(clinic_id):
    if 'user' not in session or session.get('role') != 'admin':
        return redirect(url_for('login_page'))
    clinic = VetClinic.query.get_or_404(clinic_id)
    if request.method == 'POST':
        image_file = request.files.get('image')
        if image_file and image_file.filename:
            safe_name = secure_filename(image_file.filename)
            image_filename = f"{uuid4().hex}_{safe_name}"
            image_file.save(os.path.join(app.config['UPLOAD_FOLDER'], image_filename))
            clinic.image = image_filename

        clinic.name = request.form['name']
        clinic.address = request.form['address']
        clinic.phone = request.form['phone']
        clinic.operating_hours = request.form['operating_hours']
        clinic.latitude = float(request.form['latitude']) if request.form.get('latitude') else None
        clinic.longitude = float(request.form.get('longitude')) if request.form.get('longitude') else None
        clinic.google_map_link = request.form.get('google_map_link')
        db.session.commit()
        flash('Vet clinic updated!', 'success')
        return redirect(url_for('vet_clinics'))
    return render_template('vets_clinics.html', form_mode='edit', edit_clinic=clinic, clinics=VetClinic.query.all())


@app.route('/admin/vet-clinics/delete/<int:clinic_id>', methods=['POST'])
@app.route('/vet-clinics/<int:clinic_id>/delete', methods=['POST'])
def delete_vet_clinic(clinic_id):
    if 'user' not in session or session.get('role') != 'admin':
        return redirect(url_for('login_page'))
    clinic = VetClinic.query.get_or_404(clinic_id)
    db.session.delete(clinic)
    db.session.commit()
    flash('Vet clinic deleted.', 'warning')
    return redirect(url_for('vet_clinics'))

def seed_default_users():
    defaults = [
        {"email": "admin@mmu.edu.my", "password": "admin123", "role": "admin", "username": "admin"},
        {"email": "user@student.mmu.edu.my", "password": "user1234", "role": "customer", "username": "user"},
    ]
    added_user = False
    for d in defaults:
        if not User.query.filter_by(email=d['email']).first():
            db.session.add(User(
                email    = d['email'],
                password = generate_password_hash(d['password']),
                role     = d['role']
            ))
            added_user = True
    if added_user:
        db.session.commit()

@app.route('/forgot-password')
def forgot_password_page():
    return render_template('forgot_password.html')


@app.route('/reset-password', methods=['POST'])
def reset_password():
    email = request.form.get('email').lower().strip()
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')

    user = User.query.filter_by(email=email).first()
    if not user:
        flash("Email address not found. Please register first.")
        return redirect(url_for('forgot_password_page'))

    if password != confirm_password:
        flash("Passwords do not match!")
        return redirect(url_for('forgot_password_page'))

    if not (len(password) == 8 and password.isdigit()):
        flash("Format error: Password must be exactly 8 digits!")
        return redirect(url_for('forgot_password_page'))

    user.password = generate_password_hash(password)
    db.session.commit()

    flash("Password reset successfully! Please login with your new password.")
    return redirect(url_for('show_login'))

with app.app_context():
    db.create_all()
    seed_default_users()

if __name__ == '__main__':
    app.run(debug=True)
